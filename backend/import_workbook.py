"""Import Rank + Games from local .xlsm copies. Workbooks are never written."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from catalog import SPORTS, sport_by_slug
from db import SOURCES_DIR, connect, init_db, clear_sport_data

GROUP_SHEETS = {
    "GMC",
    "State",
    "HV State",
    "G4 State",
    "G3 State",
    "G2 State",
    "G1 State",
    "NPA State",
    "NPB State",
}


def cell(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float):
        if abs(v - round(v)) < 1e-9:
            return int(round(v))
        return round(float(v), 6)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        # Normalize M/D/YYYY → ISO so chronological replay works
        if "/" in s and len(s) <= 10:
            parts = s.split("/")
            if len(parts) == 3 and parts[0].isdigit() and parts[2].isdigit():
                try:
                    m, d0, y = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return f"{y:04d}-{m:02d}-{d0:02d}"
                except ValueError:
                    pass
        return s
    return v


def _header_map(row) -> dict[str, int]:
    out = {}
    for i, v in enumerate(row):
        if v is None:
            continue
        key = str(v).strip().lower()
        if key and key not in out:
            out[key] = i
    return out


def _iter_sheet(path: Path, name: str, max_col: int = 20):
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    if name not in wb.sheetnames:
        wb.close()
        return None, []
    ws = wb[name]
    rows = ws.iter_rows(max_col=max_col, values_only=True)
    return wb, rows


def import_sport(slug: str, include_games: bool = True, *, use_sqlite: bool = False) -> dict:
    spec = sport_by_slug(slug)
    if not spec:
        raise KeyError(slug)
    path = SOURCES_DIR / spec["file"]
    if not path.exists():
        raise FileNotFoundError(path)

    from db import DB_PATH, connect_sqlite

    conn = connect_sqlite(DB_PATH) if use_sqlite else connect()
    init_db(conn)
    # Ensure tennis/hfa tables exist before clear
    try:
        from recompute import _ensure_tables

        _ensure_tables(conn)
    except Exception:
        pass
    row = conn.execute("SELECT id FROM sports WHERE slug = ?", (slug,)).fetchone()
    if row:
        sport_id = row["id"]
        conn.execute(
            "UPDATE sports SET name = ?, source_path = ? WHERE id = ?",
            (spec["name"], str(path), sport_id),
        )
    else:
        row = conn.execute(
            "INSERT INTO sports(slug, name, source_path, watch) VALUES (?, ?, ?, 1) RETURNING id",
            (slug, spec["name"], str(path)),
        ).fetchone()
        if row is None:
            # sqlite without RETURNING
            conn.execute(
                "INSERT INTO sports(slug, name, source_path, watch) VALUES (?, ?, ?, 1)",
                (slug, spec["name"], str(path)),
            )
            row = conn.execute("SELECT id FROM sports WHERE slug = ?", (slug,)).fetchone()
        sport_id = row["id"]
    clear_sport_data(conn, sport_id)

    as_of, team_count = _import_rank(conn, path, sport_id)
    game_count = 0
    if include_games:
        game_count = _import_games(conn, path, sport_id)
    _import_groups(conn, path, sport_id)

    imported_at = datetime.now().isoformat(timespec="seconds")
    mtime = datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")
    conn.execute(
        """
        UPDATE sports
        SET as_of = ?, imported_at = ?, file_mtime = ?, source_path = ?
        WHERE id = ?
        """,
        (as_of, imported_at, mtime, str(path), sport_id),
    )
    conn.commit()
    conn.close()
    return {
        "slug": slug,
        "name": spec["name"],
        "as_of": as_of,
        "teams": team_count,
        "games": game_count,
        "file": spec["file"],
    }


def _import_rank(conn, path: Path, sport_id: int) -> tuple[str | None, int]:
    wb, rows = _iter_sheet(path, "Rank")
    if rows is None:
        return None, 0
    header = next(rows, None)
    if not header:
        wb.close()
        return None, 0
    hmap = _header_map(header)
    as_of = cell(header[3]) if len(header) > 3 else None
    off_i = hmap.get("offrating")
    def_i = hmap.get("defrating")
    date_i = hmap.get("date")
    games_i = hmap.get("games")
    n_i = hmap.get("n")
    count = 0
    batch = []
    for r in rows:
        team = cell(r[3]) if len(r) > 3 else None
        rating = cell(r[4]) if len(r) > 4 else None
        if not team or not isinstance(team, str) or rating is None:
            continue
        try:
            float(rating)
        except (TypeError, ValueError):
            continue
        off = cell(r[off_i]) if off_i is not None and off_i < len(r) else None
        deff = cell(r[def_i]) if def_i is not None and def_i < len(r) else None
        last_game = cell(r[date_i]) if date_i is not None and date_i < len(r) else None
        games = cell(r[games_i]) if games_i is not None and games_i < len(r) else None
        n = cell(r[n_i]) if n_i is not None and n_i < len(r) else None
        batch.append(
            (
                sport_id,
                team,
                cell(r[2]),
                rating,
                off,
                deff,
                cell(r[1]),
                cell(r[0]),
                games,
                n,
                last_game,
            )
        )
        count += 1
        if len(batch) >= 500:
            conn.executemany(
                """
                INSERT INTO teams(sport_id, name, rank, rating, off, def, change, prev, games, n, last_game)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(sport_id, name) DO UPDATE SET
                  rank=excluded.rank, rating=excluded.rating, off=excluded.off, def=excluded.def,
                  change=excluded.change, prev=excluded.prev, games=excluded.games, n=excluded.n,
                  last_game=excluded.last_game
                """,
                batch,
            )
            batch.clear()
    if batch:
        conn.executemany(
            """
            INSERT INTO teams(sport_id, name, rank, rating, off, def, change, prev, games, n, last_game)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(sport_id, name) DO UPDATE SET
              rank=excluded.rank, rating=excluded.rating, off=excluded.off, def=excluded.def,
              change=excluded.change, prev=excluded.prev, games=excluded.games, n=excluded.n,
              last_game=excluded.last_game
            """,
            batch,
        )
    wb.close()
    return str(as_of) if as_of else None, count


def _import_games(conn, path: Path, sport_id: int) -> int:
    wb, rows = _iter_sheet(path, "Games", max_col=20)
    if rows is None:
        return 0
    header = next(rows, None)
    if not header:
        wb.close()
        return 0
    hmap = _header_map(header)
    # Tennis dual-meet line scores
    if "home team" in hmap or "match date" in hmap:
        wb.close()
        return _import_tennis_lines(conn, path, sport_id)

    def col(*names):
        for n in names:
            if n in hmap:
                return hmap[n]
        return None

    date_i = col("date")
    t1_i = col("team 1", "team1", "home team")
    s1_i = col("score 1", "score1")
    t2_i = col("team 2", "team2", "away team")
    s2_i = col("score 2", "score2")
    home_i = col("home", "h")
    par_i = col("course par", "par")
    # Fall back to classic positional layout
    if t1_i is None:
        date_i, t1_i, s1_i, t2_i, s2_i, home_i = 0, 1, 2, 3, 4, 5
    else:
        # Many Off/Def sheets label Team1/Team2 but leave Score columns unlabeled
        if s1_i is None and t1_i is not None:
            s1_i = t1_i + 1
        if s2_i is None and t2_i is not None:
            s2_i = t2_i + 1
        if home_i is None:
            home_i = (s2_i + 1) if s2_i is not None else 5

    offdef = "orioff1" in hmap or "newoff1" in hmap
    # ensure course_par column
    try:
        conn.execute("ALTER TABLE games ADD COLUMN course_par REAL")
    except Exception:
        pass

    count = 0
    batch = []
    for r in rows:
        if not r or len(r) <= max(t1_i or 0, s1_i or 0, t2_i or 0, s2_i or 0):
            continue
        t1 = cell(r[t1_i]) if t1_i is not None else None
        s1 = cell(r[s1_i]) if s1_i is not None else None
        t2 = cell(r[t2_i]) if t2_i is not None else None
        s2 = cell(r[s2_i]) if s2_i is not None else None
        if not t1 or not t2 or s1 is None or s2 is None:
            continue
        try:
            s1i, s2i = int(float(s1)), int(float(s2))
        except (TypeError, ValueError):
            continue
        h = cell(r[home_i]) if home_i is not None and home_i < len(r) else None
        par = cell(r[par_i]) if par_i is not None and par_i < len(r) else None
        try:
            par_f = float(par) if par is not None else None
        except (TypeError, ValueError):
            par_f = None
        dt = cell(r[date_i]) if date_i is not None and date_i < len(r) else None

        if offdef:
            # Prefer header names when present
            def hv(key, fallback_idx):
                i = hmap.get(key)
                if i is not None and i < len(r):
                    return cell(r[i])
                return cell(r[fallback_idx]) if fallback_idx < len(r) else None

            ori_off1 = hv("orioff1", 6)
            ori_def1 = hv("oridef1", 7)
            ori_off2 = hv("orioff2", 8)
            ori_def2 = hv("oridef2", 9)
            new_off1 = hv("newoff1", 10)
            new_def1 = hv("newdef1", 11)
            new_off2 = hv("newoff2", 12)
            new_def2 = hv("newdef2", 13)
            gd = hv("gd", 14) if "gd" in hmap or len(r) > 14 else None
            err = hv("error", 15) if "error" in hmap or len(r) > 15 else None
        else:
            # Single-rating Ori/New in rating columns
            ori_i = hmap.get("ori_team1") or hmap.get("ori1")
            ori2_i = hmap.get("ori_team2") or hmap.get("ori2")
            new_i = hmap.get("new_team1") or hmap.get("new1")
            new2_i = hmap.get("new_team2") or hmap.get("new2")
            if ori_i is None:
                ori_off1 = cell(r[6]) if len(r) > 6 else None
                ori_off2 = cell(r[7]) if len(r) > 7 else None
                new_off1 = cell(r[8]) if len(r) > 8 else None
                new_off2 = cell(r[9]) if len(r) > 9 else None
            else:
                ori_off1 = cell(r[ori_i]) if ori_i < len(r) else None
                ori_off2 = cell(r[ori2_i]) if ori2_i is not None and ori2_i < len(r) else None
                new_off1 = cell(r[new_i]) if new_i is not None and new_i < len(r) else None
                new_off2 = cell(r[new2_i]) if new2_i is not None and new2_i < len(r) else None
            ori_def1 = ori_def2 = new_def1 = new_def2 = None
            gd = cell(r[hmap["act_diff"]]) if "act_diff" in hmap else (cell(r[10]) if len(r) > 10 else None)
            err = cell(r[hmap["error"]]) if "error" in hmap else (cell(r[11]) if len(r) > 11 else None)

        batch.append(
            (
                sport_id,
                dt,
                t1,
                s1i,
                t2,
                s2i,
                1 if h == 1 else 0,
                ori_off1,
                ori_def1,
                ori_off2,
                ori_def2,
                new_off1,
                new_def1,
                new_off2,
                new_def2,
                gd,
                err,
                par_f,
            )
        )
        if len(batch) >= 2000:
            conn.executemany(
                """
                INSERT INTO games(
                  sport_id, date, team1, score1, team2, score2, home,
                  ori_off1, ori_def1, ori_off2, ori_def2,
                  new_off1, new_def1, new_off2, new_def2, gd, error, course_par
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                batch,
            )
            count += len(batch)
            batch.clear()
    if batch:
        conn.executemany(
            """
            INSERT INTO games(
              sport_id, date, team1, score1, team2, score2, home,
              ori_off1, ori_def1, ori_off2, ori_def2,
              new_off1, new_def1, new_off2, new_def2, gd, error, course_par
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )
        count += len(batch)
    wb.close()
    return count


def _import_tennis_lines(conn, path: Path, sport_id: int) -> int:
    """Import tennis dual meets into line_matches (5 positions × Game 1 scores)."""
    # Ensure table
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS line_matches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sport_id INTEGER NOT NULL,
            date TEXT,
            home_team TEXT NOT NULL,
            away_team TEXT NOT NULL,
            home INTEGER,
            s1s_h REAL, s1s_a REAL,
            s2s_h REAL, s2s_a REAL,
            s3s_h REAL, s3s_a REAL,
            s1d_h REAL, s1d_a REAL,
            s2d_h REAL, s2d_a REAL
        )
        """
    )
    wb, rows = _iter_sheet(path, "Games", max_col=20)
    if rows is None:
        return 0
    header = next(rows, None)
    if not header:
        wb.close()
        return 0
    hmap = _header_map(header)

    def idx(*names):
        for n in names:
            if n in hmap:
                return hmap[n]
        return None

    date_i = idx("match date", "date")
    home_i = idx("home team")
    away_i = idx("away team")
    home_flag_i = idx("home")
    # Game 1 score columns
    cols = {
        "s1s_h": idx("home 1st singles game 1 score"),
        "s2s_h": idx("home 2nd singles game 1 score"),
        "s3s_h": idx("home 3rd singles game 1 score"),
        "s1d_h": idx("home 1st doubles game 1 score"),
        "s2d_h": idx("home 2nd doubles game 1 score"),
        "s1s_a": idx("away 1st singles game 1 score"),
        "s2s_a": idx("away 2nd singles game 1 score"),
        "s3s_a": idx("away 3rd singles game 1 score"),
        "s1d_a": idx("away 1st doubles game 1 score"),
        "s2d_a": idx("away 2nd doubles game 1 score"),
    }
    # Positional fallback from known layout
    if home_i is None:
        date_i, home_i, away_i, home_flag_i = 0, 1, 7, 13
        cols = {
            "s1s_h": 2, "s2s_h": 3, "s3s_h": 4, "s1d_h": 5, "s2d_h": 6,
            "s1s_a": 8, "s2s_a": 9, "s3s_a": 10, "s1d_a": 11, "s2d_a": 12,
        }

    batch = []
    count = 0

    def num(r, i):
        if i is None or i >= len(r):
            return None
        v = cell(r[i])
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    for r in rows:
        if not r:
            continue
        ht = cell(r[home_i]) if home_i is not None and home_i < len(r) else None
        at = cell(r[away_i]) if away_i is not None and away_i < len(r) else None
        if not ht or not at:
            continue
        hf = cell(r[home_flag_i]) if home_flag_i is not None and home_flag_i < len(r) else 1
        batch.append(
            (
                sport_id,
                cell(r[date_i]) if date_i is not None and date_i < len(r) else None,
                ht,
                at,
                1 if hf == 1 else 0,
                num(r, cols["s1s_h"]),
                num(r, cols["s1s_a"]),
                num(r, cols["s2s_h"]),
                num(r, cols["s2s_a"]),
                num(r, cols["s3s_h"]),
                num(r, cols["s3s_a"]),
                num(r, cols["s1d_h"]),
                num(r, cols["s1d_a"]),
                num(r, cols["s2d_h"]),
                num(r, cols["s2d_a"]),
            )
        )
        if len(batch) >= 1000:
            conn.executemany(
                """
                INSERT INTO line_matches(
                  sport_id, date, home_team, away_team, home,
                  s1s_h, s1s_a, s2s_h, s2s_a, s3s_h, s3s_a, s1d_h, s1d_a, s2d_h, s2d_a
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                batch,
            )
            count += len(batch)
            batch.clear()
    if batch:
        conn.executemany(
            """
            INSERT INTO line_matches(
              sport_id, date, home_team, away_team, home,
              s1s_h, s1s_a, s2s_h, s2s_a, s3s_h, s3s_a, s1d_h, s1d_a, s2d_h, s2d_a
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )
        count += len(batch)
    wb.close()
    return count


def _import_groups(conn, path: Path, sport_id: int) -> None:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    for sheet in wb.sheetnames:
        if sheet not in GROUP_SHEETS:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(max_row=3, max_col=80, values_only=True))
        if len(rows) < 3:
            continue
        header, offs, defs = rows[0], rows[1], rows[2]
        entries = []
        for i, v in enumerate(header):
            name = cell(v)
            if not name or not isinstance(name, str):
                continue
            off = cell(offs[i]) if i < len(offs) else None
            deff = cell(defs[i]) if i < len(defs) else None
            rating = None
            if off is not None and deff is not None:
                try:
                    rating = round(float(off) + float(deff), 6)
                except (TypeError, ValueError):
                    rating = None
            elif off is not None:
                try:
                    rating = float(off)
                except (TypeError, ValueError):
                    rating = None
            entries.append((name, off, deff, rating))
        entries.sort(key=lambda x: -(x[3] if x[3] is not None else -9999))
        for i, (name, off, deff, rating) in enumerate(entries, 1):
            conn.execute(
                """
                INSERT INTO group_teams(sport_id, group_key, team, rank, off, def, rating)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(sport_id, group_key, team) DO UPDATE SET
                  rank=excluded.rank, off=excluded.off, def=excluded.def, rating=excluded.rating
                """,
                (sport_id, sheet, name, i, off, deff, rating),
            )
    wb.close()


def import_all(include_games: bool = True) -> list[dict]:
    results = []
    for spec in SPORTS:
        print(f"Import {spec['slug']} …", flush=True)
        results.append(import_sport(spec["slug"], include_games=include_games))
        print(" ", results[-1], flush=True)
    return results


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Import local spreadsheet copies (read-only).")
    parser.add_argument("--slug")
    parser.add_argument("--no-games", action="store_true")
    args = parser.parse_args()
    if args.slug:
        print(import_sport(args.slug, include_games=not args.no_games))
    else:
        import_all(include_games=not args.no_games)
