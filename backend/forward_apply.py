"""Apply the native engine only to games Excel has not calculated yet.

Flow after Drive Rank + Games import:
1. Teams start as spreadsheet Rank (matches Excel when coach is current).
2. Games with missing New* ratings are treated as pending.
3. Engine updates those games forward from current team state.

So: if Excel Rank is up to date → no pending games → rankings unchanged.
If new results are in Games but VBA hasn't run → site calculates them.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from catalog import sport_by_slug
from db import DB_PATH, SOURCES_DIR, connect, connect_sqlite, ensure_schema_extras, init_db, set_meta
from rating_engine import EngineState, HfaState, TeamState, LINE_POSITIONS


def _utcnow() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _norm_date(d: str | None) -> str | None:
    if not d:
        return d
    s = str(d).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    parts = s.replace("-", "/").split("/")
    if len(parts) == 3:
        try:
            m, d0, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100:
                y += 2000
            return f"{y:04d}-{m:02d}-{d0:02d}"
        except ValueError:
            return s
    return s


def load_hfa_from_calc(path: Path) -> HfaState | None:
    """Read Q2/Q3/R2/R3 from Calc when present."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, keep_vba=False)
        if "Calc" not in wb.sheetnames:
            wb.close()
            return None
        ws = wb["Calc"]

        def num(addr: str) -> float | None:
            v = ws[addr].value
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None

        q2, q3, r2, r3 = num("Q2"), num("Q3"), num("R2"), num("R3")
        # Tennis uses Y2/Y3 for home edge in some workbooks
        if q2 is None:
            q2, q3 = num("Y2"), num("Y3")
        if r2 is None:
            r2, r3 = num("Z2"), num("Z3")
        wb.close()
        if q3 and q3 > 0:
            return HfaState(
                q_sum=q2 or 0.0,
                q_n=q3,
                r_sum=r2 or 0.0,
                r_n=r3 if r3 and r3 > 0 else q3,
            )
    except Exception:
        return None
    return None


def save_hfa(conn, sport_id: int, hfa: HfaState, season_key: str | None = None) -> None:
    conn.execute(
        """
        INSERT INTO sport_hfa(sport_id, q_sum, q_n, r_sum, r_n, season_key, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(sport_id) DO UPDATE SET
          q_sum=excluded.q_sum, q_n=excluded.q_n,
          r_sum=excluded.r_sum, r_n=excluded.r_n,
          season_key=excluded.season_key, updated_at=excluded.updated_at
        """,
        (sport_id, hfa.q_sum, hfa.q_n, hfa.r_sum, hfa.r_n, season_key, _utcnow()),
    )


def _engine_from_db(conn, sport_id: int, spec: dict, hfa: HfaState) -> EngineState:
    engine_name = spec.get("engine") or "offdef"
    games_delta = float(spec.get("games_delta") or 0.75)
    games_cap = float(spec.get("games_cap") or 15.0)
    engine = EngineState(
        hfa=hfa,
        auto_seed_new=True,
        engine=engine_name,
        games_delta=games_delta,
        games_cap=games_cap,
    )
    rows = conn.execute(
        "SELECT name, off, def, rating, games FROM teams WHERE sport_id = ?",
        (sport_id,),
    ).fetchall()
    for r in rows:
        name = r["name"]
        games = float(r["games"] if r["games"] is not None else 2.75)
        if engine_name == "offdef":
            off = float(r["off"] or 0.0)
            deff = float(r["def"] or 0.0)
            st = TeamState(off=off, deff=deff, games=games)
        elif engine_name == "lines":
            # Overall rating only in teams today; split evenly across positions as seed.
            total = float(r["rating"] if r["rating"] is not None else (r["off"] or 0.0))
            each = total / len(LINE_POSITIONS)
            st = TeamState(
                off=total,
                deff=None,
                games=games,
                lines={p: each for p in LINE_POSITIONS},
            )
        else:
            rating = r["rating"] if r["rating"] is not None else r["off"]
            st = TeamState(off=float(rating or 0.0), deff=None, games=games)
        engine._register(name, st)
    return engine


def _write_teams(conn, sport_id: int, engine: EngineState, last_by_team: dict[str, str | None]) -> int:
    prior = {
        r["name"]: r
        for r in conn.execute(
            "SELECT name, change, prev, rank FROM teams WHERE sport_id = ?",
            (sport_id,),
        ).fetchall()
    }
    prior_l = {k.lower(): v for k, v in prior.items()}
    rows = []
    for key, st in engine.teams.items():
        name = engine.canonical.get(key, key)
        rating = st.rating
        off = st.off if st.deff is not None else rating
        deff = st.deff
        rows.append(
            {
                "name": name,
                "off": off,
                "def": deff,
                "rating": rating,
                "games": st.games,
                "n": (st.games - 2.75) / engine.games_delta if engine.games_delta else None,
                "last_game": last_by_team.get(name),
            }
        )
    rows.sort(key=lambda r: (-(r["rating"] if r["rating"] is not None else -1e9), r["name"]))
    conn.execute("DELETE FROM teams WHERE sport_id = ?", (sport_id,))
    batch = []
    for i, r in enumerate(rows, 1):
        old = prior.get(r["name"]) or prior_l.get(r["name"].lower(), {})
        batch.append(
            (
                sport_id,
                r["name"],
                i,
                r["rating"],
                r["off"],
                r["def"],
                old.get("change") if isinstance(old, dict) else None,
                old.get("prev") if isinstance(old, dict) else None,
                r["games"],
                r["n"],
                r.get("last_game"),
            )
        )
    if batch:
        conn.executemany(
            """
            INSERT INTO teams(sport_id, name, rank, rating, off, def, change, prev, games, n, last_game)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )
    return len(batch)


def apply_forward(slug: str, *, use_sqlite: bool = False) -> dict:
    """Update rankings for games that do not yet have New* ratings from Excel."""
    spec = sport_by_slug(slug)
    if not spec:
        raise ValueError(f"Unknown sport: {slug}")

    conn = connect_sqlite(DB_PATH) if use_sqlite else connect()
    init_db(conn)
    ensure_schema_extras(conn)
    try:
        sport = conn.execute("SELECT * FROM sports WHERE slug = ?", (slug,)).fetchone()
        if not sport:
            raise ValueError(f"Sport not in DB: {slug}")
        sport_id = sport["id"]
        path = SOURCES_DIR / spec["file"]
        group = spec.get("group") or "Fall"
        engine_name = spec.get("engine") or "offdef"

        hfa = load_hfa_from_calc(path) if path.exists() else None
        if hfa is None:
            row = conn.execute("SELECT * FROM sport_hfa WHERE sport_id = ?", (sport_id,)).fetchone()
            if row:
                hfa = HfaState(row["q_sum"], row["q_n"], row["r_sum"], row["r_n"])
            else:
                hfa = HfaState(0.0, 100.0, 0.0, 100.0)
        save_hfa(conn, sport_id, hfa)

        engine = _engine_from_db(conn, sport_id, spec, hfa)
        if not engine.teams:
            return {
                "ok": True,
                "slug": slug,
                "applied": 0,
                "reason": "no teams (import Rank first)",
                "source": "drive_rank",
            }

        last_by_team: dict[str, str | None] = {
            r["name"]: r["last_game"]
            for r in conn.execute(
                "SELECT name, last_game FROM teams WHERE sport_id = ?",
                (sport_id,),
            ).fetchall()
        }
        applied = 0

        if engine_name == "lines":
            # Tennis: apply duals that have scores but were never engine-touched.
            # Without per-position Rank import, forward tennis is best-effort from equal split.
            matches = conn.execute(
                """
                SELECT * FROM line_matches
                WHERE sport_id = ?
                ORDER BY date ASC, id ASC
                """,
                (sport_id,),
            ).fetchall()
            # Only apply the most recent duals after max team last_game
            max_last = max((_norm_date(d) or "" for d in last_by_team.values()), default="")
            for m in matches:
                date = _norm_date(m["date"])
                if max_last and date and date <= max_last:
                    continue
                scores = {}
                for pos, hcol, acol in [
                    ("1S", "s1s_h", "s1s_a"),
                    ("2S", "s2s_h", "s2s_a"),
                    ("3S", "s3s_h", "s3s_a"),
                    ("1D", "s1d_h", "s1d_a"),
                    ("2D", "s2d_h", "s2d_a"),
                ]:
                    hv, av = m[hcol], m[acol]
                    if hv is not None and av is not None:
                        scores[pos] = (float(hv), float(av))
                if not scores:
                    continue
                engine.process_tennis_dual(
                    date, m["home_team"], m["away_team"], bool(m["home"]), scores, group=group
                )
                last_by_team[engine.display_name(m["home_team"])] = date
                last_by_team[engine.display_name(m["away_team"])] = date
                applied += 1
        else:
            has_par = False
            if conn.dialect == "postgres":
                cols = conn.execute(
                    """
                    SELECT column_name AS name FROM information_schema.columns
                    WHERE table_schema = 'public' AND table_name = 'games'
                    """
                ).fetchall()
                has_par = any(c["name"] == "course_par" for c in cols)
            else:
                cols = conn._raw.execute("PRAGMA table_info(games)").fetchall()
                has_par = any(c["name"] == "course_par" for c in cols)

            select = """
                SELECT id, date, team1, score1, team2, score2, home,
                       new_off1, new_def1, new_off2, new_def2
            """
            if has_par:
                select += ", course_par"
            select += """
                FROM games
                WHERE sport_id = ?
                  AND new_off1 IS NULL
                ORDER BY date ASC, id ASC
            """
            pending = conn.execute(select, (sport_id,)).fetchall()
            updates = []
            for g in pending:
                date = _norm_date(g["date"])
                par = g["course_par"] if has_par else None
                result = engine.process_game(
                    date,
                    g["team1"],
                    g["score1"],
                    g["team2"],
                    g["score2"],
                    bool(g["home"]),
                    group=group,
                    course_par=par,
                )
                last_by_team[engine.display_name(g["team1"])] = date
                last_by_team[engine.display_name(g["team2"])] = date
                updates.append(
                    (
                        result.ori_off1,
                        result.ori_def1,
                        result.ori_off2,
                        result.ori_def2,
                        result.new_off1,
                        result.new_def1,
                        result.new_off2,
                        result.new_def2,
                        result.gd,
                        result.error,
                        g["id"],
                    )
                )
                applied += 1
            if updates:
                conn.executemany(
                    """
                    UPDATE games SET
                      ori_off1=?, ori_def1=?, ori_off2=?, ori_def2=?,
                      new_off1=?, new_def1=?, new_off2=?, new_def2=?,
                      gd=?, error=?
                    WHERE id=?
                    """,
                    updates,
                )

        teams_n = len(engine.teams)
        if applied:
            teams_n = _write_teams(conn, sport_id, engine, last_by_team)
            save_hfa(conn, sport_id, engine.hfa, engine.current_season)
            set_meta(conn, f"forward_{slug}_at", _utcnow())
            set_meta(conn, "rankings_source", "drive_rank_plus_forward")
        else:
            set_meta(conn, "rankings_source", "drive_rank")

        conn.commit()
        return {
            "ok": True,
            "slug": slug,
            "engine": engine_name,
            "applied": applied,
            "teams": teams_n,
            "hfa_q4": round(engine.hfa.q4, 6),
            "source": "drive_rank_plus_forward" if applied else "drive_rank",
        }
    finally:
        conn.close()
